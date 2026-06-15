"""
根据《冷藏品主数据.xlsx》回填 hq_products：
  - 仅 fill 当前为 NULL 的可空字段（不覆盖已有非空值）
  - is_new_product / is_private_label / is_returnable / allocation_unit / barcode 直接以 Excel 为准
  - 高/宽/深 cm → mm 乘 10
  - createdate(datetime) → introduced_at(date)

不会自己 commit，生成 SQL 后再用 psql 执行（便于回滚检视）。
"""
from openpyxl import load_workbook
import sys, os

XLSX = '冷藏品主数据.xlsx'
OUT = 'scripts/data/backfill-from-excel.sql'

def norm_sku(v):
    if v is None: return None
    s = str(v).strip()
    if s.isdigit() and len(s) == 7: s = '0' + s
    return s

def sql_str(v):
    if v is None: return 'NULL'
    s = str(v).replace("'", "''")
    return f"'{s}'"

def sql_num(v):
    if v is None or v == '': return 'NULL'
    try:
        return str(float(v))
    except Exception:
        return 'NULL'

def sql_int(v):
    if v is None or v == '': return 'NULL'
    try:
        return str(int(v))
    except Exception:
        return 'NULL'

def sql_bool_yn(v):
    if v is None or v == '': return 'NULL'
    s = str(v).strip()
    if s == '是': return 'TRUE'
    if s == '否': return 'FALSE'
    return 'NULL'

def sql_date(v):
    if v is None: return 'NULL'
    # datetime → date string
    return f"'{v.strftime('%Y-%m-%d')}'"

wb = load_workbook(XLSX, data_only=True)
ws = wb['Sheet1']
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
idx = {h: i for i, h in enumerate(header)}

lines = ["BEGIN;", ""]

count = 0
for r in rows[1:]:
    sku = norm_sku(r[idx['商品代码']])
    if not sku: continue
    name = r[idx['商品名称']]
    brand = r[idx['品牌']]
    spec = r[idx['规格']]
    unit = r[idx['计量单位']]
    series = r[idx['系列']]
    shelf = r[idx['保质期']]
    h_cm = r[idx['高/cm']]
    w_cm = r[idx['宽/cm']]
    d_cm = r[idx['深/cm']]
    whole = r[idx['批发价']]
    retail = r[idx['建议零售价']]
    intro = r[idx['createdate']]
    barcode = r[idx['国际代码']]
    is_new = r[idx['是否新品']]
    is_priv = r[idx['是否美宜佳自有品牌']]
    is_ret = r[idx['是否可退（退货标识）']]
    alloc = r[idx['配货单位']]

    def cm_to_mm(v):
        if v is None or v == '': return 'NULL'
        try: return str(float(v) * 10)
        except Exception: return 'NULL'

    lines.append(f"""UPDATE hq_products SET
  brand                  = COALESCE(brand, {sql_str(brand)}),
  spec                   = COALESCE(spec, {sql_str(spec)}),
  unit                   = COALESCE(unit, {sql_str(unit)}),
  series                 = COALESCE(series, {sql_str(series)}),
  shelf_life_days        = COALESCE(shelf_life_days, {sql_int(shelf)}),
  length_mm              = COALESCE(length_mm, {cm_to_mm(d_cm)}),
  width_mm               = COALESCE(width_mm, {cm_to_mm(w_cm)}),
  height_mm              = COALESCE(height_mm, {cm_to_mm(h_cm)}),
  wholesale_price        = COALESCE(wholesale_price, {sql_num(whole)}),
  suggested_retail_price = COALESCE(suggested_retail_price, {sql_num(retail)}),
  introduced_at          = COALESCE(introduced_at, {sql_date(intro)}),
  barcode                = COALESCE(barcode, {sql_str(barcode)}),
  is_returnable          = COALESCE(is_returnable, {sql_bool_yn(is_ret)}),
  allocation_unit        = COALESCE(allocation_unit, {sql_int(alloc)}),
  is_new_product         = COALESCE({sql_bool_yn(is_new)}, is_new_product),
  is_private_label       = COALESCE({sql_bool_yn(is_priv)}, is_private_label)
WHERE sku_code = '{sku}' AND deleted_at IS NULL;""")
    count += 1

lines.append("")
lines.append("-- 验收")
lines.append("""SELECT '回填后字段缺失' AS info,
  COUNT(*) FILTER (WHERE length_mm IS NULL) AS m_len,
  COUNT(*) FILTER (WHERE width_mm IS NULL) AS m_wid,
  COUNT(*) FILTER (WHERE height_mm IS NULL) AS m_hei,
  COUNT(*) FILTER (WHERE shelf_life_days IS NULL) AS m_shelf,
  COUNT(*) FILTER (WHERE wholesale_price IS NULL) AS m_whole,
  COUNT(*) FILTER (WHERE suggested_retail_price IS NULL) AS m_retail,
  COUNT(*) FILTER (WHERE series IS NULL) AS m_series,
  COUNT(*) FILTER (WHERE introduced_at IS NULL) AS m_intro,
  COUNT(*) FILTER (WHERE barcode IS NULL) AS m_barcode,
  COUNT(*) FILTER (WHERE is_returnable IS NULL) AS m_ret,
  COUNT(*) FILTER (WHERE allocation_unit IS NULL) AS m_alloc,
  COUNT(*) AS total
FROM hq_products WHERE deleted_at IS NULL;""")
lines.append("COMMIT;")

with open(OUT, 'w') as f:
    f.write('\n'.join(lines))
print(f'Wrote {count} UPDATE statements to {OUT}')
